# -*- coding: utf-8 -*-
"""RH Pro Report - Version Custom
Application KivyMD pour convertir un PDF RH en rapport Excel selon les règles spécifiques.
"""

from __future__ import annotations

import os
import re
import shutil
import tempfile
import threading
import traceback
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Optional, Sequence, Set, Tuple

DEPS_AVAILABLE = True
DEPS_ERROR = ""
try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    import pdfplumber
except ImportError as e:
    DEPS_AVAILABLE = False
    DEPS_ERROR = str(e)

APP_TITLE = "RH OCP Custom"
APP_SUBTITLE = "Extraction PDF → Excel (Règles spécifiques)"
APP_PACKAGE = "com.mycompany.rhocpcustom"
OUTPUT_FOLDER_NAME = "RH_Reports_Custom"

# Règles selon demande utilisateur
PRESENCE_CODES = {"MIS", "FC", "RPJ", "VS"}
EXCLUSION_CODES = {"RM", "RC", "PEAS", "CA", "CA-1", "CP"}

# Semaines par défaut (modifiables dans l'UI)
DEFAULT_WEEKS = {
    "s01_start": "2026-03-30", "s01_end": "2026-04-05",
    "s02_start": "2026-04-06", "s02_end": "2026-04-12",
    "s03_start": "2026-04-13", "s03_end": "2026-04-19",
    "s04_start": "2026-04-20", "s04_end": "2026-04-26",
    "s05_start": "2026-04-27", "s05_end": "2026-05-03",
}

# Imports Kivy/KivyMD protégés
KIVY_AVAILABLE = False
try:
    from kivy.clock import Clock
    from kivy.core.window import Window
    from kivy.metrics import dp
    from kivy.uix.image import Image as KivyImage
    from kivy.utils import platform
    from kivymd.app import MDApp
    from kivymd.uix.boxlayout import MDBoxLayout
    from kivymd.uix.button import MDFlatButton, MDRaisedButton
    from kivymd.uix.card import MDCard
    from kivymd.uix.dialog import MDDialog
    from kivymd.uix.label import MDLabel
    from kivymd.uix.progressbar import MDProgressBar
    from kivymd.uix.screen import MDScreen
    from kivymd.uix.scrollview import MDScrollView
    from kivymd.uix.textfield import MDTextField
    from kivymd.uix.snackbar import Snackbar

    KIVY_AVAILABLE = True
except Exception:
    Clock = None
    Window = None
    dp = lambda value: value
    platform = "unknown"
    MDApp = object

def get_output_dir() -> Path:
    candidates = [
        Path.home() / "Downloads",
        Path.home(),
        Path(tempfile.gettempdir()),
    ]
    for base in candidates:
        try:
            output_dir = base / OUTPUT_FOLDER_NAME
            output_dir.mkdir(parents=True, exist_ok=True)
            return output_dir
        except Exception:
            continue
    return Path(".")

class DataEngine:
    def __init__(self):
        self.errors = []

    def evaluate_day(self, value: str) -> int:
        if not value:
            return 0
        val = str(value).strip().upper()
        if re.match(r"^\d{1,2}[:Hh]\d{2}$", val):
            return 1
        if val in PRESENCE_CODES:
            return 1
        if val in EXCLUSION_CODES:
            return 0
        for code in PRESENCE_CODES:
            if code in val:
                return 1
        for code in EXCLUSION_CODES:
            if code in val:
                return 0
        return 0

    def process_pdf(self, pdf_path: str, weeks: Dict[str, date], progress_cb: Callable) -> str:
        employees = []
        with pdfplumber.open(pdf_path) as pdf:
            total_pages = len(pdf.pages)
            for idx, page in enumerate(pdf.pages):
                progress_cb(int((idx / total_pages) * 100), f"Traitement page {idx+1}/{total_pages}...")
                text = page.extract_text()
                if not text:
                    continue
                
                name_match = re.search(r"-\s*([^|]+?)\s+Matricule:\s*(\d+)", text)
                if not name_match:
                    continue
                name = name_match.group(1).strip()
                matricule = name_match.group(2).strip()
                
                days_data = {}
                lines = text.split('\n')
                for line in lines:
                    date_match = re.search(r"^(LUN|MAR|MER|JEU|VEN|SAM|DIM)\s+(\d{2}/\d{2}/\d{4})", line)
                    if date_match:
                        date_str = date_match.group(2)
                        current_date = datetime.strptime(date_str, "%d/%m/%Y").date()
                        rest = line[date_match.end():].strip()
                        
                        # Heuristique simple pour la présence
                        valide = 0
                        if any(c in rest for c in PRESENCE_CODES):
                            valide = 1
                        elif any(c in rest for c in EXCLUSION_CODES):
                            valide = 0
                        elif len(re.findall(r"\d{1,2}:\d{2}", rest)) >= 2:
                            valide = 1
                            
                        days_data[current_date] = valide

                # Calcul des KPIs
                def sum_week(start, end):
                    return sum(days_data.get(start + timedelta(days=i), 0) for i in range((end - start).days + 1))

                s01 = sum_week(weeks["s01_start"], weeks["s01_end"])
                s02 = sum_week(weeks["s02_start"], weeks["s02_end"])
                s03 = sum_week(weeks["s03_start"], weeks["s03_end"])
                s04 = sum_week(weeks["s04_start"], weeks["s04_end"])
                s05 = sum_week(weeks["s05_start"], weeks["s05_end"])
                
                pp = s01 + s02 + s03 + s04 + s05
                
                pf = sum(days_data.get(date(2026, 4, i), 0) for i in range(1, 31))
                
                employees.append({
                    "nom": f"{name} ({matricule})",
                    "s01": s01, "s02": s02, "s03": s03, "s04": s04, "s05": s05,
                    "pp": pp, "pf": pf
                })

        return self.generate_excel(employees)

    def generate_excel(self, employees: List[Dict]) -> str:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rapport Présence"
        
        headers = ["Nom & Prénom", "S01", "S02", "S03", "S04", "S05", "PP", "PF"]
        ws.append(headers)
        
        # Styles basiques
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="0B3D5C", end_color="0B3D5C", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            
        for emp in employees:
            ws.append([emp["nom"], emp["s01"], emp["s02"], emp["s03"], emp["s04"], emp["s05"], emp["pp"], emp["pf"]])
            
        output_path = get_output_dir() / f"Rapport_Presences_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        wb.save(str(output_path))
        return str(output_path)

if KIVY_AVAILABLE:
    class RHCustomApp(MDApp):
        def build(self):
            self.theme_cls.primary_palette = "Teal"
            self.engine = DataEngine()
            self.pdf_path = "c:\\Users\\DELL\\Desktop\\EtatMensuel (prime 04).pdf" # Par défaut pour le test
            
            screen = MDScreen()
            layout = MDBoxLayout(orientation="vertical", padding="20dp", spacing="10dp")
            
            layout.add_widget(MDLabel(text="RH OCP - Custom", font_style="H5", halign="center"))
            
            self.lbl_file = MDLabel(text=f"Fichier : {self.pdf_path}", halign="center")
            layout.add_widget(self.lbl_file)
            
            self.btn_run = MDRaisedButton(text="Lancer l'extraction", pos_hint={"center_x": 0.5})
            self.btn_run.bind(on_press=self.start_process)
            layout.add_widget(self.btn_run)
            
            self.progress = MDProgressBar(value=0)
            layout.add_widget(self.progress)
            
            self.lbl_status = MDLabel(text="Prêt", halign="center")
            layout.add_widget(self.lbl_status)
            
            screen.add_widget(layout)
            return screen

        def start_process(self, instance):
            self.btn_run.disabled = True
            threading.Thread(target=self.run_thread, daemon=True).start()

        def run_thread(self):
            weeks = {
                "s01_start": date(2026, 3, 30), "s01_end": date(2026, 4, 5),
                "s02_start": date(2026, 4, 6), "s02_end": date(2026, 4, 12),
                "s03_start": date(2026, 4, 13), "s03_end": date(2026, 4, 19),
                "s04_start": date(2026, 4, 20), "s04_end": date(2026, 4, 26),
                "s05_start": date(2026, 4, 27), "s05_end": date(2026, 5, 3),
            }
            try:
                out = self.engine.process_pdf(self.pdf_path, weeks, self.update_progress)
                Clock.schedule_once(lambda dt: self.finish(True, out))
            except Exception as e:
                Clock.schedule_once(lambda dt: self.finish(False, str(e)))

        def update_progress(self, val, msg):
            Clock.schedule_once(lambda dt: self.set_progress(val, msg))

        def set_progress(self, val, msg):
            self.progress.value = val
            self.lbl_status.text = msg

        def finish(self, success, msg):
            self.btn_run.disabled = False
            if success:
                self.lbl_status.text = f"Succès ! Fichier créé : {msg}"
            else:
                self.lbl_status.text = f"Erreur : {msg}"

    if __name__ == "__main__":
        RHCustomApp().run()
else:
    print("Kivy non dispo")
