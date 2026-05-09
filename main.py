# -*- coding: utf-8 -*-
"""RH Pro Report
Application KivyMD pour convertir un PDF RH en deux rapports Excel.

Cette version garde le moteur de traitement indépendant de l'interface afin de
pouvoir tester la génération Excel même quand Kivy n'est pas installé sur PC/CI.
"""

from __future__ import annotations

import os
import re
import shutil
import tempfile
import threading
import traceback
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Optional, Sequence, Tuple

import openpyxl
import pdfplumber
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

APP_TITLE = "RH Pro Report"
APP_SUBTITLE = "PDF → Prime de Performance + Heures Supplémentaires"
APP_PACKAGE = "org.ocp.rhproreport"
OUTPUT_FOLDER_NAME = "RH_Reports"

ABSENCE_CODES = {"RM", "RC", "PEAS", "CA", "RHJ", "IRR", "JF", "ABS", "MAL"}
PRESENCE_CODES = {"MIS", "RPJ", "FC", "RP", "P", "PR", "TR", "AP"}
HEADER_KEYWORDS = {"MATRICULE", "MLE", "NOM", "PRENOM", "PRÉNOM", "CODE", "POSTE"}

# ─────────────────────────────────────────────────────────────
#  Imports Kivy/KivyMD protégés
# ─────────────────────────────────────────────────────────────
KIVY_AVAILABLE = False
try:  # L'import est volontairement optionnel pour permettre les tests moteur.
    from kivy.clock import Clock
    from kivy.core.window import Window
    from kivy.metrics import dp
    from kivy.uix.image import Image as KivyImage
    from kivy.utils import platform
    from kivymd.app import MDApp
    from kivymd.uix.boxlayout import MDBoxLayout
    from kivymd.uix.button import MDFlatButton, MDRaisedButton
    from kivymd.uix.card import MDCard
    from kivymd.uix.dialog import MDDialog
    from kivymd.uix.label import MDLabel
    from kivymd.uix.progressbar import MDProgressBar
    from kivymd.uix.screen import MDScreen
    from kivymd.uix.scrollview import MDScrollView
    from kivymd.uix.textfield import MDTextField

    try:
        from kivymd.uix.snackbar import MDSnackbar, MDSnackbarText  # KivyMD >= 1.2
    except Exception:  # pragma: no cover - compat anciennes versions
        MDSnackbar = None
        MDSnackbarText = None
        from kivymd.uix.snackbar import Snackbar

    try:
        from plyer import filechooser, storagepath
    except Exception:  # pragma: no cover
        filechooser = None
        storagepath = None

    KIVY_AVAILABLE = True
except Exception:  # pragma: no cover - environnement sans interface graphique
    Clock = None
    Window = None
    dp = lambda value: value  # noqa: E731
    platform = "unknown"
    MDApp = object
    filechooser = None
    storagepath = None


# ─────────────────────────────────────────────────────────────
#  Utilitaires fichiers
# ─────────────────────────────────────────────────────────────
def app_base_dir() -> Path:
    """Dossier racine fiable pour les ressources incluses dans l'app."""
    return Path(__file__).resolve().parent


def get_logo_path() -> Optional[str]:
    """Cherche le logo OCP dans les emplacements probables."""
    candidates = [
        app_base_dir() / "ocp_logo.png",
        Path("ocp_logo.png"),
        Path("/data/data/org.ocp.rhproreport/files/app/ocp_logo.png"),
    ]
    for candidate in candidates:
        if candidate.exists() and candidate.is_file():
            return str(candidate)
    return None


def get_output_dir() -> Path:
    """Retourne un dossier de sortie robuste sur Android et sur ordinateur.

    Priorité : Downloads quand accessible. Sinon, dossier utilisateur/app.
    """
    candidates: List[Path] = []

    try:
        if storagepath:
            downloads = storagepath.get_downloads_dir()
            if downloads:
                candidates.append(Path(downloads))
    except Exception:
        pass

    # Sur Android, app_storage_path est toujours accessible sans permissions lourdes.
    try:
        if platform == "android":
            from android.storage import app_storage_path  # type: ignore

            candidates.append(Path(app_storage_path()))
    except Exception:
        pass

    candidates.extend([
        Path.home() / "Downloads",
        Path.home(),
        Path(tempfile.gettempdir()),
    ])

    last_error: Optional[Exception] = None
    seen: set[str] = set()
    for base in candidates:
        try:
            base = base.expanduser()
            key = str(base.resolve()) if base.exists() else str(base)
            if key in seen:
                continue
            seen.add(key)
            output_dir = base / OUTPUT_FOLDER_NAME
            output_dir.mkdir(parents=True, exist_ok=True)
            test_file = output_dir / ".write_test"
            test_file.write_text("ok", encoding="utf-8")
            test_file.unlink(missing_ok=True)
            return output_dir
        except Exception as exc:
            last_error = exc

    raise OSError(f"Impossible de créer le dossier de sortie : {last_error}")


def request_android_permissions() -> None:
    """Demande les permissions utiles quand l'app tourne sur Android."""
    if not KIVY_AVAILABLE or platform != "android":
        return
    try:
        from android.permissions import Permission, request_permissions  # type: ignore

        request_permissions([
            Permission.READ_EXTERNAL_STORAGE,
            Permission.WRITE_EXTERNAL_STORAGE,
        ])
    except Exception:
        # Les versions Android récentes peuvent utiliser le Storage Access Framework
        # via le sélecteur de fichier sans permission classique.
        pass


def materialize_pdf_path(selected_path: str) -> str:
    """Convertit une sélection fichier en chemin lisible par pdfplumber.

    - Chemin local : retourné tel quel.
    - URI Android content:// : tentative de copie vers un fichier temporaire app.
    """
    if not selected_path:
        raise ValueError("Aucun fichier PDF sélectionné.")

    selected_path = str(selected_path).strip()
    if selected_path.startswith("content://"):
        if platform != "android":
            raise ValueError("URI Android content:// reçue hors Android.")
        try:
            from android.storage import app_storage_path  # type: ignore
            from jnius import autoclass  # type: ignore

            Uri = autoclass("android.net.Uri")
            PythonActivity = autoclass("org.kivy.android.PythonActivity")
            activity = PythonActivity.mActivity
            resolver = activity.getContentResolver()
            input_stream = resolver.openInputStream(Uri.parse(selected_path))
            if input_stream is None:
                raise ValueError("Impossible d'ouvrir le fichier sélectionné.")

            out_dir = Path(app_storage_path()) / "selected_files"
            out_dir.mkdir(parents=True, exist_ok=True)
            out_path = out_dir / f"selected_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            buffer = bytearray(1024 * 64)
            with open(out_path, "wb") as output:
                while True:
                    count = input_stream.read(buffer)
                    if count == -1 or count is None:
                        break
                    if count > 0:
                        output.write(buffer[:count])
            input_stream.close()
            return str(out_path)
        except Exception as exc:
            raise ValueError(
                "Le fichier a été sélectionné via Android content:// mais n'a pas pu être copié. "
                "Essayez de le placer dans Downloads puis sélectionnez-le à nouveau. "
                f"Détail : {exc}"
            ) from exc

    p = Path(selected_path)
    if not p.exists() or not p.is_file():
        raise FileNotFoundError(f"Fichier introuvable : {selected_path}")
    if p.suffix.lower() != ".pdf":
        raise ValueError("Veuillez choisir un fichier PDF valide.")
    return str(p)


# ─────────────────────────────────────────────────────────────
#  Styles Excel communs
# ─────────────────────────────────────────────────────────────
def common_styles() -> Tuple[Border, Alignment]:
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return thin, center


def style_range_border(ws: openpyxl.worksheet.worksheet.Worksheet, cell_range: str, border: Border) -> None:
    for row in ws[cell_range]:
        for cell in row:
            cell.border = border


@dataclass
class EmployeeRecord:
    matricule: str
    code: str
    name: str
    poste: str
    s01: int
    s02: int
    s03: int
    s04: int
    s05: int
    pp: int
    pf: int
    v04: float
    v05: float
    v06: float
    v07: float
    total_supp: float

    def as_dict(self) -> Dict[str, Any]:
        return self.__dict__.copy()


# ─────────────────────────────────────────────────────────────
#  Moteur de données
# ─────────────────────────────────────────────────────────────
class DataEngine:
    """Extraction PDF + génération Excel.

    Le moteur est volontairement sans dépendance Kivy pour faciliter les tests.
    """

    def __init__(self) -> None:
        self.errors: List[str] = []
        self.report_month: Optional[Tuple[int, int]] = None

    def process_pdf(
        self,
        pdf_path: str,
        periods: Dict[str, Optional[date]],
        progress_callback: Callable[[int, str], None] = lambda value, message: None,
    ) -> Tuple[bool, str, str]:
        self.errors.clear()
        try:
            pdf_path = materialize_pdf_path(pdf_path)
            self.validate_periods(periods)
            self.report_month = self.infer_report_month(periods)

            employees_data: List[Dict[str, Any]] = []
            with pdfplumber.open(pdf_path) as pdf:
                total_pages = max(len(pdf.pages), 1)
                for page_index, page in enumerate(pdf.pages, start=1):
                    progress_callback(
                        int((page_index - 1) / total_pages * 70),
                        f"Lecture page {page_index}/{total_pages}…",
                    )
                    tables = self.extract_tables_from_page(page)
                    for table in tables:
                        for row_number, row in enumerate(table, start=1):
                            record = self.parse_row(row, periods)
                            if record:
                                employees_data.append(record.as_dict())
                            elif row and any(self.clean_cell(cell) for cell in row):
                                # Ligne non exploitable : on garde une trace légère seulement.
                                preview = " | ".join(self.clean_cell(c) for c in row[:5])[:160]
                                if preview and not self.is_header_or_noise(row):
                                    self.errors.append(f"Page {page_index}, ligne {row_number} ignorée : {preview}")
                    progress_callback(
                        int(page_index / total_pages * 75),
                        f"Analyse page {page_index}/{total_pages}…",
                    )

            employees_data = self.deduplicate_records(employees_data)
            if not employees_data:
                raise ValueError(
                    "Aucune ligne employé exploitable trouvée dans le PDF. "
                    "Vérifiez que le PDF contient un tableau texte, pas une image scannée."
                )

            progress_callback(80, "Génération du fichier Prime de Performance…")
            file_perf = self.generate_excel_performance(employees_data, periods)

            progress_callback(90, "Génération du fichier Heures Supplémentaires…")
            file_supp = self.generate_excel_heures_supp(employees_data)

            self.write_log_file()
            progress_callback(100, "Terminé avec succès !")
            return True, file_perf, file_supp

        except Exception as exc:
            self.errors.append(traceback.format_exc())
            try:
                self.write_log_file()
            except Exception:
                pass
            progress_callback(0, f"Erreur : {exc}")
            return False, str(exc), ""

    def extract_tables_from_page(self, page: Any) -> List[List[List[Any]]]:
        """Extraction tolérante avec deux stratégies pdfplumber."""
        tables: List[List[List[Any]]] = []
        strategies = [
            {},
            {
                "vertical_strategy": "text",
                "horizontal_strategy": "text",
                "snap_tolerance": 3,
                "join_tolerance": 3,
                "intersection_tolerance": 5,
            },
        ]
        for settings in strategies:
            try:
                extracted = page.extract_tables(settings) if settings else page.extract_tables()
                if extracted:
                    tables.extend(extracted)
                    break
            except Exception as exc:
                self.errors.append(f"Extraction table échouée : {exc}")
        if not tables:
            try:
                text = page.extract_text() or ""
                rows = [re.split(r"\s{2,}|\t", line.strip()) for line in text.splitlines() if line.strip()]
                if rows:
                    tables.append(rows)
            except Exception as exc:
                self.errors.append(f"Extraction texte échouée : {exc}")
        return tables

    # ── Parsing ──────────────────────────────────────────────
    def parse_row(self, row: Sequence[Any], periods: Dict[str, Optional[date]]) -> Optional[EmployeeRecord]:
        try:
            if not row or self.is_header_or_noise(row):
                return None

            cells = [self.clean_cell(cell) for cell in row]
            if len([c for c in cells if c]) < 3:
                return None

            matricule = cells[0]
            if not matricule or not re.search(r"\d", matricule):
                return None

            code = cells[1] if len(cells) > 1 else ""
            name = cells[2] if len(cells) > 2 else ""
            daily_start = self.detect_daily_start(cells)

            if daily_start <= 3:
                poste = ""
            else:
                poste = cells[3] if len(cells) > 3 else ""

            # Certains PDF placent le poste après le nom, d'autres non.
            if not name and len(cells) > 1:
                name = cells[1]

            daily_data = cells[daily_start:daily_start + 31]
            if len(daily_data) < 31:
                daily_data.extend([""] * (31 - len(daily_data)))

            v04, v05, v06, v07 = self.extract_hours(cells, daily_start)
            s01, s02, s03, s04, s05, pp, pf = self.calculate_dynamic_kpis(daily_data, periods)
            return EmployeeRecord(
                matricule=matricule,
                code=code,
                name=name,
                poste=poste,
                s01=s01,
                s02=s02,
                s03=s03,
                s04=s04,
                s05=s05,
                pp=pp,
                pf=pf,
                v04=v04,
                v05=v05,
                v06=v06,
                v07=v07,
                total_supp=round(v05 + v06 + v07, 2),
            )
        except Exception as exc:
            preview = " | ".join(self.clean_cell(c) for c in row[:8])[:180]
            self.errors.append(f"Erreur parsing ligne [{preview}] : {exc}")
            return None

    def is_header_or_noise(self, row: Sequence[Any]) -> bool:
        text = " ".join(self.clean_cell(c).upper() for c in row if self.clean_cell(c))
        if not text:
            return True
        if "SUIVI" in text and "PRIME" in text:
            return True
        if "DIRECTION" in text or "PERIODE" in text or "PÉRIODE" in text:
            return True
        first = self.clean_cell(row[0]).upper() if row else ""
        if first in {"MATRICULE", "MLE", "N°", "NO", "N"}:
            return True
        hit_count = sum(1 for word in HEADER_KEYWORDS if word in text)
        return hit_count >= 3

    def clean_cell(self, value: Any) -> str:
        if value is None:
            return ""
        return str(value).replace("\n", " ").replace("\r", " ").strip()

    def normalize_code(self, value: Any) -> str:
        val = self.clean_cell(value).upper()
        val = re.sub(r"\s+", "", val)
        return val

    def looks_like_day_value(self, value: Any) -> bool:
        val = self.normalize_code(value)
        if not val:
            return True
        if val in ABSENCE_CODES or val in PRESENCE_CODES:
            return True
        if self.safe_float(value) > 0:
            return True
        if re.fullmatch(r"\d{1,2}[:Hh]\d{1,2}", val):
            return True
        return False

    def detect_daily_start(self, cells: Sequence[str]) -> int:
        """Détecte où commencent les 31 colonnes journalières."""
        best_start = 4 if len(cells) >= 39 else 3
        best_score = -1
        max_start = min(8, max(3, len(cells) - 8))
        for start in range(3, max_start + 1):
            window = cells[start:start + 31]
            if not window:
                continue
            score = sum(1 for value in window if self.looks_like_day_value(value))
            # Bonus si après les jours il reste 3-6 colonnes numériques d'heures.
            tail = cells[start + 31:start + 37]
            score += min(4, sum(1 for value in tail if self.safe_float(value) >= 0 and self.clean_cell(value) != ""))
            if score > best_score:
                best_score = score
                best_start = start
        return best_start

    def evaluate_day(self, value: Any) -> int:
        val = self.normalize_code(value)
        if not val:
            return 0
        if val in ABSENCE_CODES:
            return 0
        if val in PRESENCE_CODES:
            return 1
        return 1 if self.safe_float(value) > 0 else 0

    def validate_periods(self, periods: Dict[str, Optional[date]]) -> None:
        if not periods.get("s01_start") or not periods.get("s01_end"):
            raise ValueError("Veuillez entrer au moins la période S01 complète.")
        for week in ["s01", "s02", "s03", "s04", "s05"]:
            start = periods.get(f"{week}_start")
            end = periods.get(f"{week}_end")
            if bool(start) != bool(end):
                raise ValueError(f"La période {week.upper()} doit avoir une date début et une date fin.")
            if start and end:
                if end < start:
                    raise ValueError(f"La date fin de {week.upper()} est avant la date début.")
                if (end - start).days > 10:
                    self.errors.append(f"Attention : {week.upper()} dépasse 10 jours ({start} → {end}).")

    def infer_report_month(self, periods: Dict[str, Optional[date]]) -> Tuple[int, int]:
        """Déduit le mois du rapport même si S01/S05 chevauchent un autre mois."""
        scores: Dict[Tuple[int, int], int] = {}
        for start_key, end_key in [(f"s{i:02d}_start", f"s{i:02d}_end") for i in range(1, 6)]:
            start = periods.get(start_key)
            end = periods.get(end_key)
            if not start or not end:
                continue
            current = start
            while current <= end:
                scores[(current.year, current.month)] = scores.get((current.year, current.month), 0) + 1
                current += timedelta(days=1)
        if not scores:
            today = date.today()
            return today.year, today.month
        return max(scores.items(), key=lambda item: (item[1], item[0]))[0]

    def calculate_dynamic_kpis(
        self,
        daily_data: Sequence[Any],
        periods: Dict[str, Optional[date]],
    ) -> Tuple[int, int, int, int, int, int, int]:
        days_status = [self.evaluate_day(d) for d in daily_data[:31]]
        if len(days_status) < 31:
            days_status.extend([0] * (31 - len(days_status)))

        report_year, report_month = self.report_month or self.infer_report_month(periods)

        def sum_period(start: Optional[date], end: Optional[date]) -> int:
            if not start or not end:
                return 0
            total = 0
            current = start
            while current <= end:
                if current.year == report_year and current.month == report_month:
                    day_index = current.day - 1
                    if 0 <= day_index < len(days_status):
                        total += days_status[day_index]
                current += timedelta(days=1)
            return total

        s01 = sum_period(periods.get("s01_start"), periods.get("s01_end"))
        s02 = sum_period(periods.get("s02_start"), periods.get("s02_end"))
        s03 = sum_period(periods.get("s03_start"), periods.get("s03_end"))
        s04 = sum_period(periods.get("s04_start"), periods.get("s04_end"))
        s05 = sum_period(periods.get("s05_start"), periods.get("s05_end"))
        pp = s01 + s02 + s03 + s04 + s05
        pf = sum(days_status)
        return s01, s02, s03, s04, s05, pp, pf

    def safe_float(self, val: Any) -> float:
        txt = self.clean_cell(val)
        if not txt:
            return 0.0
        txt = txt.replace(" ", "").replace(",", ".")
        txt = txt.replace("H", ":").replace("h", ":")
        match = re.search(r"[-+]?\d+(?:\.\d+)?(?::\d+(?:\.\d+)?)?", txt)
        if not match:
            return 0.0
        token = match.group(0)
        try:
            if ":" in token:
                hours, minutes = token.split(":", 1)
                return round(float(hours) + float(minutes) / 60.0, 2)
            return float(token)
        except Exception:
            return 0.0

    def extract_hours(self, cells: Sequence[str], daily_start: int = 4) -> Tuple[float, float, float, float]:
        # Les colonnes après les 31 jours sont plus fiables que les 4 dernières dans
        # certains PDF qui ajoutent des colonnes vides à droite.
        tail = list(cells[daily_start + 31:])
        numeric_tail = [self.safe_float(value) for value in tail if self.clean_cell(value) != ""]
        if len(numeric_tail) >= 4:
            return tuple(numeric_tail[-4:])  # type: ignore[return-value]
        fallback = list(cells[-4:]) if len(cells) >= 4 else list(cells)
        values = [self.safe_float(value) for value in fallback]
        values = ([0.0] * (4 - len(values))) + values
        return tuple(values[-4:])  # type: ignore[return-value]

    def deduplicate_records(self, employees: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Évite les doublons fréquents quand pdfplumber extrait deux fois un tableau."""
        output: List[Dict[str, Any]] = []
        seen: set[Tuple[str, str, str]] = set()
        for emp in employees:
            key = (str(emp.get("matricule", "")).strip(), str(emp.get("name", "")).strip(), str(emp.get("code", "")).strip())
            if key in seen:
                continue
            seen.add(key)
            output.append(emp)
        return output

    def write_log_file(self) -> Optional[str]:
        if not self.errors:
            return None
        log_path = get_output_dir() / "errors_log.txt"
        with open(log_path, "w", encoding="utf-8") as fh:
            fh.write(f"{APP_TITLE} — journal généré le {datetime.now():%Y-%m-%d %H:%M:%S}\n")
            fh.write("=" * 72 + "\n\n")
            for idx, error in enumerate(self.errors, start=1):
                fh.write(f"[{idx}] {error}\n\n")
        return str(log_path)

    # ─────────────────────────────────────────────────────
    #  FICHIER 1 : Prime de Performance
    # ─────────────────────────────────────────────────────
    def generate_excel_performance(self, employees_data: List[Dict[str, Any]], periods: Dict[str, Optional[date]]) -> str:
        thin, center = common_styles()
        left_align = Alignment(horizontal="left", vertical="center")
        bold_font = Font(bold=True, size=10, name="Calibri")
        normal_font = Font(bold=False, size=10, name="Calibri")
        title_fill = PatternFill(start_color="0B3D5C", end_color="0B3D5C", fill_type="solid")
        title_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        alt_fill = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")

        ap1_emps = [e for e in employees_data if str(e.get("poste", "")).strip().upper() == "AP1"]
        suivi_emps = [e for e in employees_data if str(e.get("poste", "")).strip().upper() != "AP1"]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Prime Performance"

        for col_letter, width in {
            "A": 10.54, "B": 7.09, "C": 24.63, "D": 14.09,
            "E": 12.63, "F": 8.91, "G": 10.0,
            "H": 8.54, "I": 8.09, "J": 8.09, "K": 8.09, "L": 8.09,
            "M": 8.91, "N": 8.91, "O": 18.45,
        }.items():
            ws.column_dimensions[col_letter].width = width

        s = periods.get("s01_start")
        e = periods.get("s05_end") or periods.get("s04_end") or periods.get("s03_end") or periods.get("s02_end") or periods.get("s01_end")
        periode_str = f"Période du  {s:%d/%m/%Y}  au  {e:%d/%m/%Y}" if s and e else "Période non définie"

        ws.merge_cells("A1:A3")
        logo_path = get_logo_path()
        if logo_path:
            try:
                xl_logo = XLImage(logo_path)
                xl_logo.width = 120
                xl_logo.height = 50
                ws.add_image(xl_logo, "A1")
            except Exception as exc:
                self.errors.append(f"Logo non inséré (performance) : {exc}")

        ws["B1"] = "DIRECTION AXE NORD"
        ws["B1"].font = bold_font
        ws.merge_cells("H1:N1")
        ws["H1"] = periode_str
        ws["H1"].font = normal_font
        ws["H1"].alignment = center
        ws["B2"] = "DIRECTION DE SITE DE KHOURIBGA"
        ws["B2"].font = bold_font
        ws["B3"] = "OIK/PS"
        ws["B3"].font = bold_font
        ws["O3"] = "Mois:"
        ws["O3"].font = bold_font

        def header_cell(coord: str, value: str, fill: Optional[PatternFill] = None) -> None:
            c = ws[coord]
            c.value = value
            c.font = title_font if fill else bold_font
            c.alignment = center
            c.border = thin
            if fill:
                c.fill = fill

        for coord, value in [("A5", "Matricule"), ("B5", "Code"), ("C5", "Nom /Prénom"), ("D5", "Type de poste")]:
            header_cell(coord, value, title_fill)
        ws.merge_cells("E5:G5")
        header_cell("E5", "Note", title_fill)
        ws.merge_cells("H5:I5")
        header_cell("H5", "Présence sur poste", title_fill)
        style_range_border(ws, "A5:I5", thin)
        ws.row_dimensions[5].height = 18

        for idx, emp in enumerate(ap1_emps[:5], start=6):
            values = [emp["matricule"], emp["code"], emp["name"], emp["poste"]]
            for col, value in enumerate(values, start=1):
                c = ws.cell(row=idx, column=col, value=value)
                c.font = normal_font
                c.alignment = left_align if col == 3 else center
                c.border = thin
            ws.row_dimensions[idx].height = 16

        ws.merge_cells("A13:D13")
        ws["A13"] = "SUIVI DE LA PRIME DE PERFORMANCE"
        ws["A13"].font = title_font
        ws["A13"].alignment = center
        ws["A13"].fill = title_fill
        ws.merge_cells("E13:G13")
        ws["E13"] = "Note"
        ws["E13"].font = title_font
        ws["E13"].alignment = center
        ws["E13"].fill = title_fill
        ws.merge_cells("H13:N13")
        ws["H13"] = "Présence sur poste"
        ws["H13"].font = title_font
        ws["H13"].alignment = center
        ws["H13"].fill = title_fill
        style_range_border(ws, "A13:N13", thin)
        ws.row_dimensions[13].height = 18

        suivi_headers = [
            "Matricule", "Code", "Nom /Prénom", "Type de poste",
            "Rapidité", "Qualité", "Initiative",
            "S01", "S02", "S03", "S04", "S05",
            "Nbre des jrs P.P", "Nbre des jrs P.F", "Note H",
        ]
        for col, header in enumerate(suivi_headers, start=1):
            c = ws.cell(row=14, column=col, value=header)
            c.font = bold_font
            c.alignment = center
            c.border = thin
            c.fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
        ws.row_dimensions[14].height = 30
        ws.freeze_panes = "A15"
        ws.auto_filter.ref = f"A14:O{max(14, len(suivi_emps) + 14)}"

        for idx, emp in enumerate(suivi_emps, start=15):
            row_data = [
                emp["matricule"], emp["code"], emp["name"], emp["poste"],
                None, None, None,
                emp["s01"], emp["s02"], emp["s03"], emp["s04"], emp["s05"],
                emp["pp"], emp["pf"],
            ]
            for col, value in enumerate(row_data, start=1):
                c = ws.cell(row=idx, column=col, value=value)
                c.border = thin
                c.font = normal_font
                c.alignment = left_align if col == 3 else center
                if idx % 2 == 0:
                    c.fill = alt_fill
            note = ws.cell(row=idx, column=15)
            note.value = f"=IFERROR(SUM(E{idx}:G{idx})/6,0)"
            note.border = thin
            note.font = normal_font
            note.alignment = center
            note.number_format = "0.00"
            if idx % 2 == 0:
                note.fill = alt_fill
            ws.row_dimensions[idx].height = 17

        ws.sheet_view.showGridLines = False
        path = get_output_dir() / f"Prime_Performance_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        wb.save(str(path))
        return str(path)

    # ─────────────────────────────────────────────────────
    #  FICHIER 2 : Heures Supplémentaires
    # ─────────────────────────────────────────────────────
    def generate_excel_heures_supp(self, employees_data: List[Dict[str, Any]]) -> str:
        thin, center = common_styles()
        dark_red = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
        gold_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        light_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        silver_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
        bronze_fill = PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid")
        medal_font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Heures Supp"
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:I1")
        ws["A1"] = "RAPPORT DES HEURES SUPPLÉMENTAIRES — CLASSEMENT DÉCROISSANT"
        ws["A1"].font = Font(bold=True, color="FFFFFF", size=13, name="Calibri")
        ws["A1"].fill = dark_red
        ws["A1"].alignment = center
        ws.row_dimensions[1].height = 35

        ws.merge_cells("A2:I2")
        ws["A2"] = f"Généré le : {datetime.now():%d/%m/%Y à %H:%M}  |  Direction OCP Khouribga"
        ws["A2"].font = Font(italic=True, color="8B0000", size=9, name="Calibri")
        ws["A2"].alignment = center
        ws.row_dimensions[2].height = 20

        headers = [
            "Rang", "Matricule", "Nom de l'Agent",
            "V04 (H. Normales)", "V05 (H. Supp 25%)",
            "V06 (H. Supp 50%)", "V07 (H. Supp 100%)",
            "Total H. Supp", "Remarque",
        ]
        for col, header in enumerate(headers, start=1):
            c = ws.cell(row=3, column=col, value=header)
            c.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
            c.fill = dark_red
            c.alignment = center
            c.border = thin
        ws.row_dimensions[3].height = 30
        ws.freeze_panes = "A4"

        sorted_emps = sorted(employees_data, key=lambda item: float(item.get("total_supp", 0) or 0), reverse=True)

        def remarque_style(total: float) -> Tuple[str, Font]:
            if total > 99:
                return "Alerte H.S (>99h)", Font(bold=True, color="8B0000", size=10, name="Calibri")
            if total >= 50:
                return "Surveillance", Font(bold=True, color="DAA520", size=10, name="Calibri")
            return "Normal", Font(color="228B22", size=10, name="Calibri")

        for rank, emp in enumerate(sorted_emps, start=1):
            row = rank + 3
            if rank == 1:
                rank_label, rank_fill, rank_font, row_fill = "1", gold_fill, Font(bold=True, color="000000", size=12, name="Calibri"), PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
            elif rank == 2:
                rank_label, rank_fill, rank_font, row_fill = "2", silver_fill, medal_font, PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            elif rank == 3:
                rank_label, rank_fill, rank_font, row_fill = "3", bronze_fill, medal_font, PatternFill(start_color="FBE9E7", end_color="FBE9E7", fill_type="solid")
            else:
                rank_label = str(rank)
                rank_fill = light_gray if rank % 2 == 0 else white_fill
                rank_font = Font(bold=True, size=10, name="Calibri")
                row_fill = rank_fill

            total = float(emp.get("total_supp", 0) or 0)
            rem_text, rem_font = remarque_style(total)
            row_data = [
                rank_label,
                emp.get("matricule", ""), emp.get("name", ""),
                emp.get("v04", 0), emp.get("v05", 0), emp.get("v06", 0), emp.get("v07", 0),
                total, rem_text,
            ]
            for col, value in enumerate(row_data, start=1):
                c = ws.cell(row=row, column=col, value=value)
                c.fill = row_fill
                c.alignment = center
                c.border = thin
                c.font = Font(size=10, name="Calibri")
                if col == 1:
                    c.fill = rank_fill
                    c.font = rank_font
                elif col in {4, 5, 6, 7, 8}:
                    c.number_format = "0.00"
                if col == 8:
                    c.fill = gold_fill
                    c.font = Font(bold=True, size=11, name="Calibri")
                if col == 9:
                    c.font = rem_font

        for i, width in enumerate([8, 14, 30, 18, 18, 18, 18, 20, 24], start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
        ws.auto_filter.ref = f"A3:I{max(3, len(sorted_emps) + 3)}"

        logo_path = get_logo_path()
        if logo_path:
            try:
                xl_logo = XLImage(logo_path)
                xl_logo.width = 120
                xl_logo.height = 40
                ws.add_image(xl_logo, "H1")
            except Exception as exc:
                self.errors.append(f"Logo non inséré (heures supp) : {exc}")

        path = get_output_dir() / f"Heures_Supplementaires_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        wb.save(str(path))
        return str(path)


# ─────────────────────────────────────────────────────────────
#  Interface KivyMD
# ─────────────────────────────────────────────────────────────
if KIVY_AVAILABLE:
    if platform not in {"android", "ios"}:
        Window.size = (390, 720)

    class RHApp(MDApp):
        def build(self):
            request_android_permissions()
            self.theme_cls.primary_palette = "Blue"
            self.theme_cls.theme_style = "Light"
            self.engine = DataEngine()
            self.pdf_path = ""
            self.dialog = None
            self.generated_files: Tuple[str, str] = ("", "")

            screen = MDScreen(md_bg_color=(0.96, 0.98, 1, 1))
            scroll = MDScrollView()
            layout = MDBoxLayout(
                orientation="vertical",
                padding=("18dp", "18dp", "18dp", "24dp"),
                spacing="14dp",
                size_hint_y=None,
            )
            layout.bind(minimum_height=layout.setter("height"))

            hero = MDCard(
                orientation="vertical",
                padding="18dp",
                spacing="8dp",
                radius=[22, 22, 22, 22],
                elevation=3,
                size_hint_y=None,
                height="190dp",
                md_bg_color=(1, 1, 1, 1),
            )
            logo_path = get_logo_path()
            if logo_path:
                hero.add_widget(KivyImage(source=logo_path, size_hint_y=None, height="58dp", allow_stretch=True, keep_ratio=True))
            hero.add_widget(MDLabel(text=APP_TITLE, font_style="H5", bold=True, halign="center", size_hint_y=None, height="34dp"))
            hero.add_widget(MDLabel(text=APP_SUBTITLE, font_style="Subtitle2", halign="center", theme_text_color="Secondary", size_hint_y=None, height="28dp"))
            hero.add_widget(MDLabel(text="Rapports RH propres, rapides et prêts pour Excel", font_style="Caption", halign="center", theme_text_color="Secondary", size_hint_y=None, height="24dp"))
            layout.add_widget(hero)

            dates_card = MDCard(
                orientation="vertical",
                padding="14dp",
                spacing="10dp",
                radius=[18, 18, 18, 18],
                elevation=2,
                size_hint_y=None,
                md_bg_color=(1, 1, 1, 1),
            )
            dates_card.bind(minimum_height=dates_card.setter("height"))
            dates_card.add_widget(MDLabel(text="1) Périodes de présence", font_style="Subtitle1", bold=True, size_hint_y=None, height="30dp"))
            dates_card.add_widget(MDLabel(text="Format : AAAA-MM-JJ — remplissez S01, puis ajoutez S02→S05 selon le mois.", font_style="Caption", theme_text_color="Secondary", size_hint_y=None, height="32dp"))

            self.entries: Dict[str, Any] = {}
            for week in ["S01", "S02", "S03", "S04", "S05"]:
                row = MDBoxLayout(spacing="8dp", size_hint_y=None, height="58dp")
                for suffix, label in [("_start", "Du"), ("_end", "Au")]:
                    field = MDTextField(
                        hint_text=f"{week} {label}",
                        helper_text="AAAA-MM-JJ",
                        mode="rectangle",
                        font_size="13sp",
                    )
                    self.entries[f"{week.lower()}{suffix}"] = field
                    row.add_widget(field)
                dates_card.add_widget(row)
            layout.add_widget(dates_card)

            file_card = MDCard(
                orientation="vertical",
                padding="14dp",
                spacing="12dp",
                radius=[18, 18, 18, 18],
                elevation=2,
                size_hint_y=None,
                height="210dp",
                md_bg_color=(1, 1, 1, 1),
            )
            file_card.add_widget(MDLabel(text="2) PDF source", font_style="Subtitle1", bold=True, size_hint_y=None, height="30dp"))
            self.btn_select = MDRaisedButton(text="Choisir le fichier PDF", size_hint_y=None, height="48dp")
            self.btn_select.bind(on_press=self.select_file)
            file_card.add_widget(self.btn_select)
            self.lbl_path = MDLabel(text="Aucun fichier sélectionné", halign="center", size_hint_y=None, height="34dp", theme_text_color="Secondary")
            file_card.add_widget(self.lbl_path)
            self.btn_process = MDRaisedButton(
                text="Générer les 2 fichiers Excel",
                size_hint_y=None,
                height="50dp",
                md_bg_color=(0.07, 0.42, 0.25, 1),
                disabled=True,
            )
            self.btn_process.bind(on_press=self.start_processing)
            file_card.add_widget(self.btn_process)
            layout.add_widget(file_card)

            status_card = MDCard(
                orientation="vertical",
                padding="14dp",
                spacing="10dp",
                radius=[18, 18, 18, 18],
                elevation=1,
                size_hint_y=None,
                height="116dp",
                md_bg_color=(1, 1, 1, 1),
            )
            status_card.add_widget(MDLabel(text="État du traitement", font_style="Subtitle1", bold=True, size_hint_y=None, height="28dp"))
            self.progress_bar = MDProgressBar(value=0, size_hint_y=None, height="8dp")
            status_card.add_widget(self.progress_bar)
            self.lbl_status = MDLabel(text="Prêt.", halign="center", size_hint_y=None, height="42dp")
            status_card.add_widget(self.lbl_status)
            layout.add_widget(status_card)

            scroll.add_widget(layout)
            screen.add_widget(scroll)
            return screen

        def show_snackbar(self, text: str) -> None:
            try:
                if "MDSnackbar" in globals() and MDSnackbar and MDSnackbarText:
                    MDSnackbar(MDSnackbarText(text=text), duration=3, pos_hint={"center_x": 0.5}, size_hint_x=0.92).open()
                else:
                    Snackbar(text=text).open()  # type: ignore[name-defined]
            except Exception:
                self.lbl_status.text = text

        def select_file(self, instance) -> None:
            if filechooser:
                filechooser.open_file(
                    title="Choisir un fichier PDF",
                    filters=[("PDF files", "*.pdf")],
                    on_selection=self.handle_file_selection,
                )
            else:
                try:
                    from tkinter import filedialog

                    path = filedialog.askopenfilename(filetypes=[("PDF Files", "*.pdf")])
                    if path:
                        self.handle_file_selection([path])
                except Exception as exc:
                    self.show_dialog("Erreur", f"Sélecteur de fichier indisponible : {exc}")

        def handle_file_selection(self, selection: Sequence[str]) -> None:
            if not selection:
                return
            selected = str(selection[0])
            try:
                # Validation légère ici. Le moteur refera la matérialisation au moment du traitement.
                if not selected.startswith("content://") and Path(selected).suffix.lower() != ".pdf":
                    raise ValueError("Veuillez choisir un fichier PDF.")
                self.pdf_path = selected
                self.lbl_path.text = Path(selected).name if not selected.startswith("content://") else "PDF sélectionné"
                self.btn_process.disabled = False
                self.show_snackbar("Fichier PDF sélectionné")
            except Exception as exc:
                self.show_dialog("Erreur", str(exc))

        def collect_periods(self) -> Dict[str, Optional[date]]:
            periods: Dict[str, Optional[date]] = {}
            for key, field in self.entries.items():
                txt = field.text.strip()
                periods[key] = datetime.strptime(txt, "%Y-%m-%d").date() if txt else None
            self.engine.validate_periods(periods)
            return periods

        def start_processing(self, instance) -> None:
            try:
                if not self.pdf_path:
                    self.show_snackbar("Choisissez d'abord un PDF")
                    return
                periods = self.collect_periods()
                self.btn_process.disabled = True
                self.btn_select.disabled = True
                self.progress_bar.value = 0
                self.lbl_status.text = "Démarrage…"
                threading.Thread(target=self.run_engine, args=(periods,), daemon=True).start()
            except ValueError as exc:
                self.show_dialog("Dates invalides", f"{exc}\n\nUtilisez le format AAAA-MM-JJ.")
            except Exception as exc:
                self.show_dialog("Erreur", str(exc))

        def run_engine(self, periods: Dict[str, Optional[date]]) -> None:
            def cb(value: int, message: str) -> None:
                Clock.schedule_once(lambda dt: self.update_ui(value, message))

            result = self.engine.process_pdf(self.pdf_path, periods, cb)
            Clock.schedule_once(lambda dt: self.processing_finished(result))

        def update_ui(self, value: int, message: str) -> None:
            self.progress_bar.value = value
            self.lbl_status.text = message

        def processing_finished(self, result: Tuple[bool, str, str]) -> None:
            self.btn_process.disabled = False
            self.btn_select.disabled = False
            if result[0]:
                _, file_perf, file_supp = result
                self.generated_files = (file_perf, file_supp)
                self.show_dialog(
                    "Succès",
                    "2 fichiers générés avec succès.\n\n"
                    f"Dossier : {get_output_dir()}\n\n"
                    f"• {Path(file_perf).name}\n"
                    f"• {Path(file_supp).name}",
                )
            else:
                self.show_dialog("Erreur", result[1])

        def show_dialog(self, title: str, text: str) -> None:
            if self.dialog:
                self.dialog.dismiss()
            self.dialog = MDDialog(
                title=title,
                text=text,
                buttons=[MDFlatButton(text="OK", on_press=lambda x: self.dialog.dismiss())],
            )
            self.dialog.open()

else:
    class RHApp:  # type: ignore[no-redef]
        def run(self) -> None:
            print(
                "Kivy/KivyMD n'est pas installé dans cet environnement.\n"
                "Installez les dépendances puis lancez : python main.py\n"
                "Pour tester le moteur sans interface : python smoke_test.py"
            )


if __name__ == "__main__":
    RHApp().run()
