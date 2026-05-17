# -*- coding: utf-8 -*-
"""
Moteur d'extraction spécifique pour les règles de l'utilisateur.
"""
from datetime import date, datetime, timedelta
import re
from pathlib import Path
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Règles
PRESENCE_CODES = {"MIS", "FC", "RPJ", "VS"}
EXCLUSION_CODES = {"RM", "RC", "PEAS", "CA", "CA-1", "CP"}

# Semaines
WEEKS = {
    "S01": (date(2026, 3, 30), date(2026, 4, 5)),
    "S02": (date(2026, 4, 6), date(2026, 4, 12)),
    "S03": (date(2026, 4, 13), date(2026, 4, 19)),
    "S04": (date(2026, 4, 20), date(2026, 4, 26)),
    "S05": (date(2026, 4, 27), date(2026, 5, 3)),
}

def evaluate_day(value):
    if not value:
        return 0
    val = str(value).strip().upper()
    
    if re.match(r"^\d{1,2}[:Hh]\d{2}$", val):
        return 1
        
    if val in PRESENCE_CODES:
        return 1
        
    if val in EXCLUSION_CODES:
        return 0
        
    # Vérifier si c'est une absence avec heures (ex: FC 9:00 H)
    for code in PRESENCE_CODES:
        if code in val:
            return 1
    for code in EXCLUSION_CODES:
        if code in val:
            return 0
            
    return 0

def extract_data(pdf_path):
    employees = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue
                
            # Extraire le nom et matricule
            name_match = re.search(r"-\s*([^|]+?)\s+Matricule:\s*(\d+)", text)
            if not name_match:
                continue
            name = name_match.group(1).strip()
            matricule = name_match.group(2).strip()
            
            # Extraire les lignes de tableau
            lines = text.split('\n')
            days_data = {}
            
            for line in lines:
                # Chercher une ligne qui commence par un jour et une date
                date_match = re.search(r"^(LUN|MAR|MER|JEU|VEN|SAM|DIM)\s+(\d{2}/\d{2}/\d{4})", line)
                if date_match:
                    day_str = date_match.group(2)
                    current_date = datetime.strptime(day_str, "%d/%m/%Y").date()
                    
                    # Extraire le pointage valide ou l'absence
                    # On cherche après la date
                    rest = line[date_match.end():].strip()
                    
                    # On simplifie : on cherche des codes ou des heures
                    # C'est une heuristique, l'idéal est d'analyser la structure
                    # Mais avec le format fourni dans le screenshot, on peut chercher
                    # ce qui se trouve dans la zone "POINTAGES VALIDES" ou "ABSENCES"
                    
                    valide = 0
                    # Exemple: si la ligne contient "FC 9:00 H" ou "VS 9:00 H"
                    if "FC" in rest or "VS" in rest or "MIS" in rest or "RPJ" in rest:
                        valide = 1
                    elif "RM" in rest or "RC" in rest or "CA" in rest or "CP" in rest:
                        valide = 0
                    else:
                        # Chercher si des heures sont renseignées dans POINTAGES VALIDES
                        # Dans l'OCR, on a des motifs comme "22:00 7:00"
                        times = re.findall(r"\d{1,2}:\d{2}", rest)
                        if len(times) >= 2: # Entrée et Sortie
                            valide = 1
                            
                    days_data[current_date] = valide
                    
            # Calculer les KPIs pour cet employé
            s01 = sum(days_data.get(d, 0) for d in [WEEKS["S01"][0] + timedelta(days=i) for i in range(7)])
            s02 = sum(days_data.get(d, 0) for d in [WEEKS["S02"][0] + timedelta(days=i) for i in range(7)])
            s03 = sum(days_data.get(d, 0) for d in [WEEKS["S03"][0] + timedelta(days=i) for i in range(7)])
            s04 = sum(days_data.get(d, 0) for d in [WEEKS["S04"][0] + timedelta(days=i) for i in range(7)])
            s05 = sum(days_data.get(d, 0) for d in [WEEKS["S05"][0] + timedelta(days=i) for i in range(7)])
            
            pp = s01 + s02 + s03 + s04 + s05
            
            # PF : du 01/04 au 30/04
            pf = 0
            current = date(2026, 4, 1)
            while current <= date(2026, 4, 30):
                pf += days_data.get(current, 0)
                current += timedelta(days=1)
                
            employees.append({
                "nom": f"{name} ({matricule})",
                "s01": s01,
                "s02": s02,
                "s03": s03,
                "s04": s04,
                "s05": s05,
                "pp": pp,
                "pf": pf
            })
            
    return employees

def generate_excel(employees, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rapport Présence"
    
    # Style
    header_fill = PatternFill(start_color="0B3D5C", end_color="0B3D5C", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center")
    
    headers = ["Nom & Prénom", "S01", "S02", "S03", "S04", "S05", "PP", "PF"]
    ws.append(headers)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        
    for emp in employees:
        ws.append([emp["nom"], emp["s01"], emp["s02"], emp["s03"], emp["s04"], emp["s05"], emp["pp"], emp["pf"]])
        
    wb.save(output_path)
    print(f"Fichier généré : {output_path}")

if __name__ == "__main__":
    pdf = r"c:\Users\DELL\Desktop\EtatMensuel (prime 04).pdf"
    output = r"c:\Users\DELL\Desktop\Rapport_Presences.xlsx"
    print("Extraction en cours...")
    data = extract_data(pdf)
    print(f"Extraction terminée. {len(data)} collaborateurs trouvés.")
    generate_excel(data, output)
