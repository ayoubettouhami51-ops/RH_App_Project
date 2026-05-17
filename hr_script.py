import re
import pandas as pd
import pdfplumber
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from rapidfuzz import fuzz, process

# ==========================================
# 1. Configuration & Rules
# ==========================================

# Dates and Weeks Configuration
WEEKS_CONFIG = {
    "S01": (datetime(2026, 3, 30), datetime(2026, 4, 5)),
    "S02": (datetime(2026, 4, 6), datetime(2026, 4, 12)),
    "S03": (datetime(2026, 4, 13), datetime(2026, 4, 19)),
    "S04": (datetime(2026, 4, 20), datetime(2026, 4, 26)),
    "S05": (datetime(2026, 4, 27), datetime(2026, 5, 3))
}

# Auto Grant specifically ends on 30/04 for S05 according to prompt
AUTO_WEEKS_CONFIG = WEEKS_CONFIG.copy()
AUTO_WEEKS_CONFIG["S05"] = (datetime(2026, 4, 27), datetime(2026, 4, 30))

APRIL_START = datetime(2026, 4, 1)
APRIL_END = datetime(2026, 4, 30)

# Presence/Exclusion Rules
# Présence (=1): HH:MM, MIS (Mission), RPJ (Repos Journée), FC (Formation), VS (Visite Systématique)
PERF_PRESENCE = ["MIS", "RPJ", "FC", "VS"]
# Exclusion (=0): RM, RC, PEAS, PEA, CA, CA-1, CP
PERF_EXCLUSION = ["RM", "RC", "PEAS", "PEA", "CA", "CA-1", "CP"]

AUTO_EXCLUSION = ["MIS", "FC", "RM", "RC", "PEAS", "PEA", "CA", "CA-1", "CP"]

# ==========================================
# 2. Data Extraction & Processing
# ==========================================

def is_date_in_range(date_obj, start_date, end_date):
    return start_date <= date_obj <= end_date

def parse_hours(val_str):
    if ':' in val_str:
        h, m = val_str.split(':')
        return int(h) + int(m)/60.0
    try:
        return float(val_str)
    except:
        return 0.0

def process_pdf(pdf_path):
    employees = {}
    
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue
            
            # --- Extract Matricule and Name ---
            matricule = None
            raw_name = "Inconnu"
            
            mat_match = re.search(r"Matricule\s*:\s*(\d+)", text, re.IGNORECASE)
            if mat_match:
                matricule = mat_match.group(1)
            else:
                fallback_mat = re.findall(r"\b(\d{5,6})\b", text)
                if fallback_mat:
                    matricule = fallback_mat[0]
            
            if not matricule:
                continue
                
            name_match = re.search(r"-\s*([^|]+?)\s+Matricule", text, re.IGNORECASE)
            if name_match:
                raw_name = name_match.group(1).strip()
            else:
                alt_name = re.search(r"Nom[^:]*:\s*([^\n]+)", text, re.IGNORECASE)
                if alt_name:
                    raw_name = alt_name.group(1).strip()
            
            # Initialize Employee Data if not exists
            if matricule not in employees:
                employees[matricule] = {
                    "Matricule": matricule,
                    "Name": raw_name,
                    "Perf_Attendance": {},
                    "Auto_Attendance": {},
                    "Hours": {"V04": 0.0, "V05": 0.0, "V06": 0.0, "V07": 0.0},
                    "Warnings": []
                }
            
            emp = employees[matricule]
            
            # --- Extract Daily Data ---
            lines = text.split('\n')
            for line in lines:
                date_match = re.search(r"(LUN|MAR|MER|JEU|VEN|SAM|DIM|lun|mar|mer|jeu|ven|sam|dim)\s+(\d{2}/\d{2}/\d{4})", line)
                if date_match:
                    date_str = date_match.group(2)
                    current_date = datetime.strptime(date_str, "%d/%m/%Y")
                    rest_of_line = line[date_match.end():].strip().upper()
                    
                    has_timestamp = bool(re.search(r"\d{1,2}:\d{2}", rest_of_line))
                    
                    # 1. Performance Logic
                    if any(exc in rest_of_line for exc in PERF_EXCLUSION):
                        emp["Perf_Attendance"][current_date] = 0
                    elif has_timestamp or any(pre in rest_of_line for pre in PERF_PRESENCE):
                        emp["Perf_Attendance"][current_date] = 1
                    else:
                        emp["Perf_Attendance"][current_date] = 0
                        
                    # 2. Automobile Logic
                    if any(exc in rest_of_line for exc in AUTO_EXCLUSION):
                        emp["Auto_Attendance"][current_date] = 0
                    elif has_timestamp:
                        emp["Auto_Attendance"][current_date] = 1
                    else:
                        emp["Auto_Attendance"][current_date] = 0
                        
                    # 3. Hours Fallback (Line by Line)
                    hours_matches = re.findall(r"(\d+:\d+|\d+\.\d+)", rest_of_line)
                    if len(hours_matches) >= 4:
                        emp["Hours"]["V04"] += parse_hours(hours_matches[-4])
                        emp["Hours"]["V05"] += parse_hours(hours_matches[-3])
                        emp["Hours"]["V06"] += parse_hours(hours_matches[-2])
                        emp["Hours"]["V07"] += parse_hours(hours_matches[-1])

            # Global Search for V04-V07 totals at bottom of page
            for v_key in ["V04", "V05", "V06", "V07", "V4", "V5", "V6", "V7"]:
                global_matches = re.findall(v_key + r"[\s:=]+(\d+[:.]\d+|\d+)", text, re.IGNORECASE)
                if global_matches:
                    normalized_key = v_key if len(v_key) == 3 else "V0" + v_key[1]
                    parsed_val = parse_hours(global_matches[-1])
                    if parsed_val > emp["Hours"][normalized_key]:
                        emp["Hours"][normalized_key] = parsed_val

    # Final Aggregation
    results_perf = []
    results_hours = []
    results_auto = []
    
    for mat, emp in employees.items():
        # Performance KPIs
        perf_s = {}
        for week, (start, end) in WEEKS_CONFIG.items():
            perf_s[week] = sum(1 for d, val in emp["Perf_Attendance"].items() if val == 1 and is_date_in_range(d, start, end))
        
        perf_pp = sum(perf_s.values())
        perf_pf = sum(1 for d, val in emp["Perf_Attendance"].items() if val == 1 and is_date_in_range(d, APRIL_START, APRIL_END))
        
        results_perf.append({
            "Matricule": emp["Matricule"], "Code": "", "Nom / Prénom": emp["Name"], "Type de poste": "", 
            "Rapidité": "", "Qualité": "", "Initiative": "", 
            "S01": perf_s["S01"], "S02": perf_s["S02"], "S03": perf_s["S03"], "S04": perf_s["S04"], "S05": perf_s["S05"], 
            "Nbre des jrs P.P": perf_pp, "Nbre des jrs P.F": perf_pf, "Note H": " / ".join(emp["Warnings"])
        })
        
        # Hours KPIs
        v04 = round(emp["Hours"]["V04"], 2)
        v05 = round(emp["Hours"]["V05"], 2)
        v06 = round(emp["Hours"]["V06"], 2)
        v07 = round(emp["Hours"]["V07"], 2)
        total_supp = v05 + v06 + v07
        
        results_hours.append({
            "Matricule": emp["Matricule"], "Nom de l'Agent": emp["Name"], 
            "V04 (H. Normales)": v04, "V05 (H. Supp 25%)": v05, 
            "V06 (H. Supp 50%)": v06, "V07 (H. Supp 100%)": v07, 
            "Total Heures Supplémentaires (V05+V06+V07)": total_supp, 
            "Remarque": " / ".join(emp["Warnings"])
        })
        
        # Auto KPIs
        auto_s = {}
        for week, (start, end) in AUTO_WEEKS_CONFIG.items():
            auto_s[week] = sum(1 for d, val in emp["Auto_Attendance"].items() if val == 1 and is_date_in_range(d, start, end))
        
        auto_pp = sum(auto_s.values())
        
        results_auto.append({
            "Nom & Prénom": emp["Name"], "Matricule": emp["Matricule"], 
            "S01": auto_s["S01"], "S02": auto_s["S02"], "S03": auto_s["S03"], "S04": auto_s["S04"], "S05": auto_s["S05"], 
            "PP": auto_pp
        })

    # Sort Hours Descending
    results_hours = sorted(results_hours, key=lambda x: x["Total Heures Supplémentaires (V05+V06+V07)"], reverse=True)
    
    return pd.DataFrame(results_perf), pd.DataFrame(results_hours), pd.DataFrame(results_auto)

# ==========================================
# 3. Excel Generation & Styling
# ==========================================

def style_excel(writer, df, sheet_name):
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    worksheet = writer.sheets[sheet_name]
    
    # Styling variables
    header_fill = PatternFill(start_color="0B3D5C", end_color="0B3D5C", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Apply to headers
    for col_num, value in enumerate(df.columns.values):
        cell = worksheet.cell(row=1, column=col_num+1)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        # Auto-adjust width
        worksheet.column_dimensions[cell.column_letter].width = max(len(str(value)) + 5, 12)
        
    # Apply to data
    for row in worksheet.iter_rows(min_row=2, max_row=len(df)+1, min_col=1, max_col=len(df.columns)):
        for cell in row:
            cell.alignment = center_align
            cell.border = thin_border

def generate_excel(pdf_path, output_path):
    df_perf, df_hours, df_auto = process_pdf(pdf_path)
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        style_excel(writer, df_perf, "Prime de Performance")
        style_excel(writer, df_hours, "Heures Supplémentaires")
        style_excel(writer, df_auto, "Prime Automobile")
        
    return df_perf, df_hours, df_auto

if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        pdf_file = sys.argv[1]
        out_file = "Rapport_RH_OCP_Final.xlsx"
        df_perf, df_hours, df_auto = generate_excel(pdf_file, out_file)
        print(f"✅ Fichier généré avec succès : {out_file}")
    else:
        print("Veuillez fournir le chemin du fichier PDF. Exemple: python hr_script.py Etat_Mensuel.pdf")
